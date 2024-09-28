from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
import os.path
from geopy.distance import geodesic
import smtplib
from email.mime.text import MIMEText
app = Flask(__name__)

# Define the required location (e.g., school location)
##required_location = (40.7128, -74.0060)  # Example coordinates for New York City
required_location = (19.197815, 72.8270384)
# Get the current working directory
current_directory = os.getcwd()

# Specify the absolute path to store the Excel file
excel_file = os.path.join(current_directory, 'user_locations.xlsx')

# Email configuration settings
EMAIL_HOST = 'smtp.gmail.com'
EMAIL_PORT = 587  # Use 465 for SSL
EMAIL_USER = 'fn1308062@gmail.com'
EMAIL_PASSWORD = 'vxil pwhj hqcf oamm'  # Use App Password if two-factor authentication is enabled

def send_email(recipient, name, status):
    subject = 'Child Attendance Notification'
    if status == 'present':
        body = f'Hello,\n\nYour child, {name}, is present today.\n\nBest regards,\nAtharva College Of Engineering'
    else:
        body = f'Hello,\n\nYour child, {name}, is absent today.\n\nPlease contact the school for further details.\n\nBest regards,\nAtharva College Of Engineering'

    msg = MIMEText(body)
    msg['From'] = EMAIL_USER
    msg['To'] = recipient
    msg['Subject'] = subject

    server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT)
    server.starttls()
    server.login(EMAIL_USER, EMAIL_PASSWORD)
    server.sendmail(EMAIL_USER, recipient, msg.as_string())
    server.quit()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']
    latitude = float(request.form['latitude'])
    longitude = float(request.form['longitude'])

    # Calculate the distance between the submitted location and the required location
    student_location = (latitude, longitude)
    distance = geodesic(student_location, required_location).kilometers

    # Determine the status based on the distance
    if distance <= 1.0:
        status = 'present'
    else:
        status = 'absent'

    # Load existing workbook
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Latitude', 'Longitude', 'Status'])
        wb.save(excel_file)

    # Load existing workbook
    wb = load_workbook(excel_file)
    ws = wb.active

    # Append new entry
    ws.append([name, email, latitude, longitude, status])

    # Save workbook
    wb.save(excel_file)

    # Send email with customized message
    send_email(email, name, status)

    return jsonify({'status': 'success', 'message': 'Location and status saved successfully!', 'distance': distance})

if __name__ == '__main__':
    app.run(debug=True)
